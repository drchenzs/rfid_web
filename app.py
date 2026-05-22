from flask import Flask, render_template, request, redirect, url_for, flash
import os
from datetime import datetime

# 创建Flask应用
app = Flask(__name__)

# 配置数据库
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///rfid_system_new.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'your-secret-key-here'

# 从db.py导入SQLAlchemy实例
from db import db

# 初始化数据库
db.init_app(app)

# 导入生成器
from generators.barcode_generator import BarcodeGenerator
from generators.rfid_generator import RFIDGenerator

# 初始化生成器
barcode_generator = BarcodeGenerator()
rfid_generator = RFIDGenerator()

# 延迟导入模型，避免循环依赖
class Models:
    Product = None
    Barcode = None
    RFID = None
    Category = None
    SubCategory = None
    Color = None

def init_models():
    global Models
    from models.product import Product
    from models.barcode import Barcode
    from models.rfid import RFID
    from models.category import Category, SubCategory
    from models.color import Color
    Models.Product = Product
    Models.Barcode = Barcode
    Models.RFID = RFID
    Models.Category = Category
    Models.SubCategory = SubCategory
    Models.Color = Color

# 主页路由
@app.route('/')
def index():
    # 统计数据
    product_count = Models.Product.query.count()
    barcode_count = Models.Barcode.query.count()
    
    return render_template('index.html', 
                           product_count=product_count, 
                           barcode_count=barcode_count)

# 数据管理路由（合并产品和条形码管理）
@app.route('/data_management')
def data_management():
    search_query = request.args.get('search', '')
    
    if search_query:
        # 搜索款号、颜色或尺码
        products = Models.Product.query.filter(
            (Models.Product.style_no.contains(search_query)) |
            (Models.Product.color.contains(search_query)) |
            (Models.Product.size.contains(search_query))
        ).all()
    else:
        # 获取所有产品
        products = Models.Product.query.all()
    
    return render_template('data_management.html', products=products, search_query=search_query)

# 产品管理路由（保留旧路由，重定向到新路由）
@app.route('/products')
def products():
    return redirect(url_for('data_management'))

# 条形码管理路由（保留旧路由，重定向到新路由）
@app.route('/barcodes')
def barcodes():
    return redirect(url_for('data_management'))

# RFID管理路由（暂时保留，后续修改）
@app.route('/rfids')
def rfids():
    rfids = Models.RFID.query.all()
    return render_template('rfids.html', rfids=rfids)

# 编辑产品路由
@app.route('/edit_product/<int:id>', methods=['GET', 'POST'])
def edit_product(id):
    # 获取产品
    product = Models.Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # 更新产品信息
            product.product_type = request.form['product_type']
            product.style_no = request.form['style_no']
            product.style_name = request.form.get('style_name', '')
            product.category = request.form['category']
            product.sub_category = request.form['sub_category']
            product.gender = request.form['gender']
            product.color = request.form['color']
            product.quantity = int(request.form['quantity'])
            
            # 根据产品类型更新尺码
            if product.product_type == '鞋子':
                product.size_us = request.form['size_us']
                product.size_uk = request.form.get('size_uk', '')
                product.size_eu = request.form.get('size_eu', '')
                product.size = product.size_us  # 使用US尺码作为基础尺码
            else:
                product.size = request.form['size']
                product.size_us = None
                product.size_uk = None
                product.size_eu = None
            
            # 更新配色数量和索引
            # 1. 查询当前款号的所有颜色
            # 生成基于（性别+颜色）的唯一键列表，按出现顺序排序
            # 这样不同性别的相同颜色会获得不同的索引
            gender_color_pairs = []
            for p in existing_products:
                pair = (p.gender, p.color)
                if pair not in gender_color_pairs:
                    gender_color_pairs.append(pair)
            
            # 添加当前的性别+颜色组合
            current_pair = (product.gender, product.color)
            if current_pair not in gender_color_pairs:
                gender_color_pairs.append(current_pair)
            
            # 计算当前颜色的索引（基于性别+颜色组合）
            product.color_index = gender_color_pairs.index(current_pair) + 1
            
            # 配色数量为性别+颜色组合的总数
            product.color_count = len(gender_color_pairs)
            
            # 更新该款号所有产品的配色数量
            Models.Product.query.filter_by(style_no=product.style_no).update({'color_count': product.color_count})
            
            # 提交变更
            db.session.commit()
            flash('产品更新成功！', 'success')
            return redirect(url_for('data_management'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'danger')
    
    return render_template('edit_product.html', product=product)

# 删除产品路由
@app.route('/delete_product/<int:id>')
def delete_product(id):
    # 获取产品
    product = Models.Product.query.get_or_404(id)
    
    try:
        # 删除产品
        db.session.delete(product)
        db.session.commit()
        flash('产品删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'danger')
    
    return redirect(url_for('data_management'))

# 重新生成EAN13码路由
@app.route('/regenerate_ean13/<int:id>')
def regenerate_ean13(id):
    from generators.barcode_generator import BarcodeGenerator
    
    try:
        # 获取产品
        product = Models.Product.query.get_or_404(id)
        
        # 创建条形码生成器实例
        barcode_generator = BarcodeGenerator()
        
        # 获取所有现有产品，用于计算颜色索引
        existing_products = Models.Product.query.filter_by(style_no=product.style_no).all()
        
        # 生成基于（性别+颜色）的唯一键列表，按出现顺序排序
        gender_color_pairs = []
        for p in existing_products:
            pair = (p.gender, p.color)
            if pair not in gender_color_pairs:
                gender_color_pairs.append(pair)
        
        # 计算当前产品的颜色索引
        current_pair = (product.gender, product.color)
        color_index_val = gender_color_pairs.index(current_pair) + 1
        
        # 配色数量为性别+颜色组合的总数
        color_count = len(gender_color_pairs)
        
        # 生成唯一的EAN13码
        max_attempts = 100
        attempts = 0
        ean13 = None
        
        while attempts < max_attempts:
            attempts += 1
            temp_ean13 = barcode_generator.generate_ean13(
                product.product_type, product.style_no, color_count, color_index_val, product.size
            )
            
            # 检查EAN13码是否已存在
            existing_barcode = Models.Barcode.query.filter_by(ean13=temp_ean13).first()
            if not existing_barcode:
                ean13 = temp_ean13
                break
        
        if not ean13:
            flash('无法生成唯一的EAN13码，请稍后重试！', 'danger')
            return redirect(url_for('data_management'))
        
        # 查询颜色代码
        color_obj = Models.Color.query.filter_by(name=product.color).first()
        color_code = color_obj.code if color_obj else None
        
        # 生成新的SKU
        sku = barcode_generator.generate_sku(
            product.category, product.gender, product.sub_category, product.size, product.color, color_code
        )

        # 更新或创建条形码记录
        if product.barcodes:
            # 更新现有条形码
            barcode = product.barcodes[0]
            barcode.ean13 = ean13
            barcode.sku = sku
        else:
            # 创建新条形码
            barcode = Models.Barcode(
                product_id=product.id,
                ean13=ean13,
                sku=sku
            )
            db.session.add(barcode)
        
        # 更新产品的配色数量和索引
        product.color_count = color_count
        product.color_index = color_index_val
        
        # 更新该款号所有产品的配色数量
        Models.Product.query.filter_by(style_no=product.style_no).update({'color_count': color_count})
        
        # 提交变更
        db.session.commit()
        flash('EAN13码已成功重新生成！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'重新生成失败：{str(e)}', 'danger')
    
    return redirect(url_for('data_management'))

# 生成条形码路由
@app.route('/generate_barcode', methods=['GET', 'POST'])
def generate_barcode():
    # 处理GET请求，获取所有类别数据
    categories = Models.Category.query.order_by(Models.Category.name).all()
    
    if request.method == 'POST':
        try:
            # 初始化重复记录列表
            duplicates = []
            
            # 获取表单数据
            product_type = request.form['product_type']
            style_no = request.form['style_no']
            style_name = request.form.get('style_name', '')  # 款名，可为空
            category = request.form['category']
            sub_category = request.form['sub_category']
            
            # 获取性别列表
            genders = request.form.getlist('genders[]')
            
            # 处理表单数据中的列表格式
            def get_nested_list(prefix):
                nested_dict = {}
                
                # 调试：显示所有表单数据
                print(f"All form data: {list(request.form.items())}")
                
                # 获取所有以prefix开头的表单字段
                for key in request.form.keys():
                    if key.startswith(prefix):
                        print(f"Processing key: {key}")
                        
                        # 使用正则表达式提取索引
                        import re
                        match = re.match(r'{}\[(\d+)\](?:\[(\d+)\])?(?:\[\])?'.format(prefix), key)
                        if not match:
                            continue
                        
                        groups = match.groups()
                        gender_idx = int(groups[0])
                        color_idx = int(groups[1]) if groups[1] else None
                        
                        # 获取值列表
                        values = request.form.getlist(key)
                        # 过滤掉空值
                        values = [v.strip() for v in values if v.strip()]
                        if not values:
                            continue
                        
                        try:
                            if prefix == 'colors':
                                # 颜色字段：colors[gender_index][] 或 colors[gender_index]
                                if gender_idx not in nested_dict:
                                    nested_dict[gender_idx] = {}
                                # 对于颜色，每个值是一个单独的颜色
                                for idx, color in enumerate(values):
                                    nested_dict[gender_idx][idx] = [color]
                            elif prefix == 'sizes' or prefix.startswith('shoe_sizes_'):
                                # 尺码字段：sizes[gender_index][color_index][] 或 sizes[gender_index][color_index]
                                if color_idx is not None:
                                    if gender_idx not in nested_dict:
                                        nested_dict[gender_idx] = {}
                                    nested_dict[gender_idx][color_idx] = values
                        except ValueError as e:
                            # 忽略无法转换为整数的情况
                            print(f"ValueError processing key {key}: {e}")
                            continue
                
                print(f"Result for {prefix}: {nested_dict}")
                return nested_dict
            
            # 调试：检查性别列表
            print(f"Genders: {genders}")
            
            # 获取颜色、尺码和鞋子尺码的嵌套列表
            colors_dict = get_nested_list('colors')
            sizes_dict = get_nested_list('sizes')
            shoe_sizes_us_dict = get_nested_list('shoe_sizes_us')
            shoe_sizes_uk_dict = get_nested_list('shoe_sizes_uk')
            shoe_sizes_eu_dict = get_nested_list('shoe_sizes_eu')
            
            # 调试：检查获取的数据
            print(f"Colors dict: {colors_dict}")
            print(f"Sizes dict: {sizes_dict}")
            print(f"Shoe sizes US dict: {shoe_sizes_us_dict}")
            
            # 处理每个性别
            print(f"Processing {len(genders)} genders")
            
            # 检查是否有数据要处理
            product_created = False
            
            # 添加安全限制，避免死循环
            max_genders = 10
            max_colors = 10
            max_sizes = 20
            
            for gender_index, gender in enumerate(genders):
                # 限制处理的性别数量，避免死循环
                if gender_index >= max_genders:
                    print(f"Exceeded maximum number of genders to process: {max_genders}")
                    break
                    
                # 跳过空的性别值
                if not gender.strip():
                    continue
                    
                print(f"Processing gender {gender} at index {gender_index}")
                # 获取该性别的所有颜色
                if gender_index in colors_dict:
                    color_list = colors_dict[gender_index]
                    print(f"Processing {len(color_list)} colors for gender {gender}")
                    
                    # 处理每个颜色
                    color_count = 0
                    for color_index, color_values in color_list.items():
                        # 限制处理的颜色数量，避免死循环
                        color_count += 1
                        if color_count > max_colors:
                            print(f"Exceeded maximum number of colors to process: {max_colors}")
                            break
                        
                        for color in color_values:
                            # 调试：显示当前处理的颜色和尺码
                            print(f"Processing color {color} at index {color_index} for gender {gender}")
                            print(f"Sizes dict for gender {gender_index}, color {color_index}: {sizes_dict.get(gender_index, {}).get(color_index, 'Not found')}")
                            print(f"Shoe sizes US dict for gender {gender_index}, color {color_index}: {shoe_sizes_us_dict.get(gender_index, {}).get(color_index, 'Not found')}")
                            
                            # 获取该颜色的所有尺码
                            if product_type == '鞋子':
                                # 处理鞋子尺码
                                if gender_index in shoe_sizes_us_dict and color_index in shoe_sizes_us_dict[gender_index]:
                                    us_sizes = shoe_sizes_us_dict[gender_index][color_index]
                                    uk_sizes = shoe_sizes_uk_dict.get(gender_index, {}).get(color_index, [''] * len(us_sizes))
                                    eu_sizes = shoe_sizes_eu_dict.get(gender_index, {}).get(color_index, [''] * len(us_sizes))
                                    
                                    # 遍历每个尺码组合
                                    size_count = 0
                                    for size_us, size_uk, size_eu in zip(us_sizes, uk_sizes, eu_sizes):
                                        # 限制处理的尺码数量，避免死循环
                                        size_count += 1
                                        if size_count > max_sizes:
                                            print(f"Exceeded maximum number of sizes to process: {max_sizes}")
                                            break
                                        
                                        if size_us:  # 只处理有US尺码的记录
                                            # 自动计算配色数量和索引
                                            # 获取所有现有产品
                                            existing_products = Models.Product.query.filter_by(style_no=style_no).all()
                                            
                                            # 生成基于（性别+颜色）的唯一键列表，按出现顺序排序
                                            # 这样不同性别的相同颜色会获得不同的索引
                                            gender_color_pairs = []
                                            for p in existing_products:
                                                pair = (p.gender, p.color)
                                                if pair not in gender_color_pairs:
                                                    gender_color_pairs.append(pair)
                                            
                                            # 添加当前的性别+颜色组合
                                            current_pair = (gender, color)
                                            if current_pair not in gender_color_pairs:
                                                gender_color_pairs.append(current_pair)
                                            
                                            # 计算当前颜色的索引（基于性别+颜色组合）
                                            color_index_val = gender_color_pairs.index(current_pair) + 1
                                            
                                            # 配色数量为性别+颜色组合的总数
                                            color_count = len(gender_color_pairs)
                                            
                                            # 使用US尺码作为基础尺码
                                            size = size_us
                                            
                                            # 生成唯一的EAN13码，确保不存在于数据库中
                                            max_attempts = 100
                                            attempts = 0
                                            ean13 = None
                                            
                                            while attempts < max_attempts:
                                                attempts += 1
                                                temp_ean13 = barcode_generator.generate_ean13(
                                                    product_type, style_no, color_count, color_index_val, size
                                                )
                                                
                                                # 检查EAN13码是否已存在
                                                existing_barcode = Models.Barcode.query.filter_by(ean13=temp_ean13).first()
                                                if not existing_barcode:
                                                    ean13 = temp_ean13
                                                    break
                                            
                                            if not ean13:
                                                raise Exception(f"Failed to generate unique EAN13 code after {max_attempts} attempts")
                                            
                                            # 查询颜色代码
                                            color_obj = Models.Color.query.filter_by(name=color).first()
                                            color_code = color_obj.code if color_obj else None
                                            
                                            # 生成SKU
                                            sku = barcode_generator.generate_sku(
                                                category, gender, sub_category, size, color, color_code
                                            )
                                            
                                            # 创建产品记录
                                            product = Models.Product(
                                                product_type=product_type,
                                                style_no=style_no,
                                                style_name=style_name,
                                                category=category,
                                                sub_category=sub_category,
                                                gender=gender,
                                                color=color,
                                                size=size,
                                                size_us=size_us,
                                                size_uk=size_uk,
                                                size_eu=size_eu,
                                                color_count=color_count,
                                                color_index=color_index_val,
                                                quantity=1  # 默认数量为1
                                            )
                                            db.session.add(product)
                                            # 先提交产品记录，确保product.id有值
                                            db.session.commit()
                                            
                                            # 创建条形码记录
                                            barcode = Models.Barcode(
                                                product_id=product.id,
                                                ean13=ean13,
                                                sku=sku
                                            )
                                            db.session.add(barcode)
                                            
                                            # 更新该款号所有产品的配色数量
                                            Models.Product.query.filter_by(style_no=style_no).update({'color_count': color_count})
                                            
                                            # 设置产品创建标志
                                            product_created = True
                                else:
                                    # 调试：显示为什么没有处理鞋子尺码
                                    print(f"No shoe sizes found for gender {gender_index}, color {color_index}")
                            else:
                                # 处理普通尺码
                                print(f"Processing as regular product type: {product_type}")
                                
                                # 检查是否存在对应的尺码
                                if gender_index in sizes_dict and color_index in sizes_dict[gender_index]:
                                    sizes = sizes_dict[gender_index][color_index]
                                    
                                    # 遍历每个尺码
                                    size_count = 0
                                    for size in sizes:
                                        # 限制处理的尺码数量，避免死循环
                                        size_count += 1
                                        if size_count > max_sizes:
                                            print(f"Exceeded maximum number of sizes to process: {max_sizes}")
                                            break
                                        
                                        if size:  # 只处理有尺码的记录
                                            # 自动计算配色数量和索引
                                            # 获取所有现有产品
                                            existing_products = Models.Product.query.filter_by(style_no=style_no).all()
                                            
                                            # 生成基于（性别+颜色）的唯一键列表，按出现顺序排序
                                            # 这样不同性别的相同颜色会获得不同的索引
                                            gender_color_pairs = []
                                            for p in existing_products:
                                                pair = (p.gender, p.color)
                                                if pair not in gender_color_pairs:
                                                    gender_color_pairs.append(pair)
                                            
                                            # 添加当前的性别+颜色组合
                                            current_pair = (gender, color)
                                            if current_pair not in gender_color_pairs:
                                                gender_color_pairs.append(current_pair)
                                            
                                            # 计算当前颜色的索引（基于性别+颜色组合）
                                            color_index_val = gender_color_pairs.index(current_pair) + 1
                                            
                                            # 配色数量为性别+颜色组合的总数
                                            color_count = len(gender_color_pairs)
                                            
                                            # 生成唯一的EAN13码，确保不存在于数据库中
                                            max_attempts = 100
                                            attempts = 0
                                            ean13 = None
                                            
                                            while attempts < max_attempts:
                                                attempts += 1
                                                temp_ean13 = barcode_generator.generate_ean13(
                                                    product_type, style_no, color_count, color_index_val, size
                                                )
                                                
                                                # 检查EAN13码是否已存在
                                                existing_barcode = Models.Barcode.query.filter_by(ean13=temp_ean13).first()
                                                if not existing_barcode:
                                                    ean13 = temp_ean13
                                                    break
                                            
                                            if not ean13:
                                                raise Exception(f"Failed to generate unique EAN13 code after {max_attempts} attempts")
                                            
                                            # 查询颜色代码
                                            color_obj = Models.Color.query.filter_by(name=color).first()
                                            color_code = color_obj.code if color_obj else None
                                            
                                            # 生成SKU
                                            sku = barcode_generator.generate_sku(
                                                category, gender, sub_category, size, color, color_code
                                            )
                                            
                                            # 调试：显示创建产品的信息
                                            print(f"Creating product: style_no={style_no}, gender={gender}, color={color}, size={size}")
                                            
                                            # 创建产品记录
                                            product = Models.Product(
                                                product_type=product_type,
                                                style_no=style_no,
                                                style_name=style_name,
                                                category=category,
                                                sub_category=sub_category,
                                                gender=gender,
                                                color=color,
                                                size=size,
                                                color_count=color_count,
                                                color_index=color_index_val,
                                                quantity=1  # 默认数量为1
                                            )
                                            db.session.add(product)
                                            # 先提交产品记录，确保product.id有值
                                            print("Committing product...")
                                            db.session.commit()
                                            print(f"Product committed with id: {product.id}")
                                            
                                            # 创建条形码记录
                                            print(f"Creating barcode: ean13={ean13}, sku={sku}")
                                            barcode = Models.Barcode(
                                                product_id=product.id,
                                                ean13=ean13,
                                                sku=sku
                                            )
                                            db.session.add(barcode)
                                            print("Barcode added to session")
                                            
                                            # 更新该款号所有产品的配色数量
                                            Models.Product.query.filter_by(style_no=style_no).update({'color_count': color_count})
                                            
                                            # 设置产品创建标志
                                            product_created = True
                                else:
                                    # 调试：显示为什么没有处理普通尺码
                                    print(f"No regular sizes found for gender {gender_index}, color {color_index}")
                                    
                                    # 即使没有尺码，也尝试创建一个产品记录（使用默认尺码）
                                    print("Creating product with default size")
                                    
                                    # 自动计算配色数量和索引
                                    # 获取所有现有产品
                                    existing_products = Models.Product.query.filter_by(style_no=style_no).all()
                                    
                                    # 生成基于（性别+颜色）的唯一键列表，按出现顺序排序
                                    # 这样不同性别的相同颜色会获得不同的索引
                                    gender_color_pairs = []
                                    for p in existing_products:
                                        pair = (p.gender, p.color)
                                        if pair not in gender_color_pairs:
                                            gender_color_pairs.append(pair)
                                    
                                    # 添加当前的性别+颜色组合
                                    current_pair = (gender, color)
                                    if current_pair not in gender_color_pairs:
                                        gender_color_pairs.append(current_pair)
                                    
                                    # 计算当前颜色的索引（基于性别+颜色组合）
                                    color_index_val = gender_color_pairs.index(current_pair) + 1
                                    
                                    # 配色数量为性别+颜色组合的总数
                                    color_count_val = len(gender_color_pairs)
                                    
                                    # 生成唯一的EAN13码，确保不存在于数据库中
                                    max_attempts = 100
                                    attempts = 0
                                    ean13 = None
                                    
                                    while attempts < max_attempts:
                                        attempts += 1
                                        temp_ean13 = barcode_generator.generate_ean13(
                                            product_type, style_no, color_count_val, color_index_val, "Default"
                                        )
                                        
                                        # 检查EAN13码是否已存在
                                        existing_barcode = Models.Barcode.query.filter_by(ean13=temp_ean13).first()
                                        if not existing_barcode:
                                            ean13 = temp_ean13
                                            break
                                    
                                    if not ean13:
                                        raise Exception(f"Failed to generate unique EAN13 code after {max_attempts} attempts")
                                    
                                    # 查询颜色代码
                                    color_obj = Models.Color.query.filter_by(name=color).first()
                                    color_code = color_obj.code if color_obj else None
                                    
                                    # 生成SKU
                                    sku = barcode_generator.generate_sku(
                                        category, gender, sub_category, "Default", color, color_code
                                    )
                                    
                                    # 创建产品记录
                                    product = Models.Product(
                                        product_type=product_type,
                                        style_no=style_no,
                                        style_name=style_name,
                                        category=category,
                                        sub_category=sub_category,
                                        gender=gender,
                                        color=color,
                                        size="Default",
                                        color_count=color_count_val,
                                        color_index=color_index_val,
                                        quantity=1  # 默认数量为1
                                    )
                                    db.session.add(product)
                                    # 先提交产品记录，确保product.id有值
                                    db.session.commit()
                                    
                                    # 创建条形码记录
                                    barcode = Models.Barcode(
                                        product_id=product.id,
                                        ean13=ean13,
                                        sku=sku
                                    )
                                    db.session.add(barcode)
                                    
                                    # 更新该款号所有产品的配色数量
                                    Models.Product.query.filter_by(style_no=style_no).update({'color_count': color_count_val})
                                    
                                    # 设置产品创建标志
                                    product_created = True
            
            # 提交所有变更
            db.session.commit()
            
            # 显示结果信息
            if product_created:
                if duplicates:
                    # 如果有重复记录，显示详细的重复信息
                    flash(f'部分产品生成成功！发现 {len(duplicates)} 条重复记录，已跳过：', 'warning')
                    for duplicate in duplicates:
                        flash(duplicate, 'info')
                else:
                    # 如果没有重复记录，显示成功信息
                    flash('条形码批量生成成功！', 'success')
            else:
                # 如果没有创建任何产品，显示错误信息
                flash('未生成任何产品，请检查输入的性别、颜色和尺码数据！', 'danger')
                print('No products were created!')
                print(f'Genders: {genders}')
                print(f'Colors dict: {colors_dict}')
                print(f'Sizes dict: {sizes_dict}')
                print(f'Shoe sizes US dict: {shoe_sizes_us_dict}')
            
        except Exception as e:
            db.session.rollback()
            flash(f'生成失败：{str(e)}', 'danger')
    
    return render_template('generate_barcode.html', barcode_result=None, categories=categories)

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import os
import io
from flask import send_file

# 全局变量，用于存储生成的RFID数据，方便下载
generated_rfid_data = []

# 保存上传的Excel文件，用于后续处理
import tempfile
import os
from werkzeug.utils import secure_filename

# 生成RFID路由
@app.route('/generate_rfid', methods=['GET', 'POST'])
def generate_rfid():
    rfid_results = []
    global generated_rfid_data
    generated_rfid_data = []
    
    if request.method == 'POST':
        try:
            # 检查是否是第一步（上传文件）还是第二步（确认映射）
            step = request.form.get('step')
            
            if step == 'generate':
                # 第二步：用户确认映射后，生成RFID标签
                file_path = request.form.get('file_path')
                barcode_col_index = int(request.form.get('barcode_col'))
                qty_col_index = int(request.form.get('qty_col'))
                
                if not file_path or not os.path.exists(file_path):
                    flash('文件不存在，请重新上传', 'danger')
                    return render_template('generate_rfid.html', rfid_results=rfid_results)
                
                # 读取Excel文件
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 获取表头
                headers = [cell.value for cell in ws[1]]
                
                # 创建结果列表
                rfid_results_list = []
                
                # 遍历数据行，生成RFID
                for row in ws.iter_rows(min_row=2):
                    # 获取完整行数据
                    row_data = [cell.value for cell in row]
                    
                    # 获取EAN13码和数量
                    ean13 = row[barcode_col_index].value
                    qty = row[qty_col_index].value
                    
                    if ean13 and qty:
                        try:
                            qty = int(qty)
                            if qty > 0:
                                # 直接生成多个RFID标签，不检查数据库
                                rfids = rfid_generator.generate_multiple_rfids(str(ean13), qty)
                                
                                # 保存到结果列表，包含完整行数据
                                for rfid in rfids:
                                    rfid_results.append(rfid)
                                    # 保存完整行数据和生成的RFID
                                    rfid_results_list.append({
                                        'row_data': row_data,
                                        'ean13': str(ean13),
                                        'rfid': rfid
                                    })
                            
                        except ValueError as e:
                            flash(f'处理行时出错: {e}\n行数据: {[cell.value for cell in row]}', 'danger')
                            continue
                
                # 保存生成的RFID数据，用于后续下载
                generated_rfid_data = {
                    'headers': headers,
                    'data': rfid_results_list
                }
                
                # 删除临时文件
                if os.path.exists(file_path):
                    os.remove(file_path)
                
                flash(f'成功生成 {len(rfid_results)} 个RFID标签！', 'success')
                return render_template('generate_rfid.html', rfid_results=rfid_results)
            else:
                # 第一步：上传文件，显示列映射确认页面
                if 'excel_file' not in request.files:
                    flash('请上传Excel文件', 'danger')
                    return render_template('generate_rfid.html', rfid_results=rfid_results)
                
                file = request.files['excel_file']
                
                if file.filename == '':
                    flash('请选择Excel文件', 'danger')
                    return render_template('generate_rfid.html', rfid_results=rfid_results)
                
                # 保存临时文件
                temp_dir = tempfile.gettempdir()
                filename = secure_filename(file.filename)
                file_path = os.path.join(temp_dir, filename)
                file.save(file_path)
                
                # 读取Excel文件
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 获取表头
                headers = [cell.value if cell.value else f'列{i+1}' for i, cell in enumerate(ws[1])]
                
                # 模糊匹配列名
                def fuzzy_match_column(headers, target_keywords):
                    for i, header in enumerate(headers):
                        if header:
                            header_lower = str(header).lower()
                            for keyword in target_keywords:
                                if keyword.lower() in header_lower:
                                    return i
                    return None
                
                # 定义关键词列表
                barcode_keywords = ['barcode', '条码', '条形码', 'ean', 'ean13']
                qty_keywords = ['qty', 'quantity', '数量', '数']
                
                # 匹配列
                default_barcode_col = fuzzy_match_column(headers, barcode_keywords)
                default_qty_col = fuzzy_match_column(headers, qty_keywords)
                
                # 获取前5行数据作为预览
                preview_data = []
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6)):
                    row_data = [cell.value for cell in row]
                    preview_data.append(row_data)
                
                # 显示列映射确认页面
                return render_template('map_columns.html', 
                                     columns=headers,
                                     default_barcode_col=default_barcode_col,
                                     default_qty_col=default_qty_col,
                                     preview_data=preview_data,
                                     file_path=file_path)
            
        except Exception as e:
            # 清理临时文件
            if 'file_path' in locals() and os.path.exists(file_path):
                os.remove(file_path)
            flash(f'处理失败：{str(e)}', 'danger')
            return render_template('generate_rfid.html', rfid_results=rfid_results)
    
    return render_template('generate_rfid.html', rfid_results=rfid_results)

# 下载RFID Excel文件路由
@app.route('/download_rfid_excel')
def download_rfid_excel():
    global generated_rfid_data
    
    if not generated_rfid_data or 'headers' not in generated_rfid_data:
        flash('没有生成RFID数据，请先生成RFID标签', 'warning')
        return redirect(url_for('generate_rfid'))
    
    # 创建新Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "RFID生成结果"
    
    # 获取原文件表头，并添加RFID列
    headers = generated_rfid_data['headers'] + ["RFID"]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入数据
    for item in generated_rfid_data['data']:
        # 获取原行数据，并添加生成的RFID
        row_data = item['row_data'] + [item['rfid']]
        ws.append(row_data)
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # 将Excel文件保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="rfid_results.xlsx", as_attachment=True)

# 从选中行生成RFID路由
@app.route('/generate_rfid_from_selected', methods=['POST'])
def generate_rfid_from_selected():
    """从选中的产品行生成RFID标签"""
    try:
        import json
        from flask import jsonify
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import io
        
        # 获取请求数据
        data = request.get_json()
        products = data.get('products', [])
        
        if not products:
            return jsonify({'success': False, 'error': '请选择要生成RFID的数据行'})
        
        total_rfids = 0
        
        # 收集生成的RFID数据，用于导出Excel
        rfid_data = []
        
        # 遍历选中的产品
        for item in products:
            product_id = item.get('product_id')
            quantity = item.get('quantity', 0)
            
            if not product_id or not isinstance(quantity, int) or quantity <= 0:
                continue
            
            # 获取产品信息
            product = Models.Product.query.get_or_404(product_id)
            
            # 获取产品的EAN13码和SKU
            if not product.barcodes:
                continue
            
            barcode = product.barcodes[0]
            ean13 = barcode.ean13
            sku = barcode.sku
            
            # 为每个产品单独维护流水号计数器，从1开始
            for product_serial in range(1, quantity + 1):
                # 生成RFID标签
                rfid_code = rfid_generator.generate_rfid(ean13, quantity, product_serial)
                
                # 不写入数据库，只生成RFID用于导出
                
                # 收集产品信息和生成的RFID，用于导出Excel
                rfid_data.append({
                    '产品ID': product.id,
                    '产品类型': product.product_type,
                    '款号': product.style_no,
                    '款名': product.style_name or '',
                    '大类别': product.category,
                    '小类别': product.sub_category,
                    '性别': product.gender,
                    '颜色': product.color,
                    '基础尺码': product.size,
                    'US尺码': product.size_us or '',
                    'UK尺码': product.size_uk or '',
                    'EU尺码': product.size_eu or '',
                    '数量': product.quantity,
                    'EAN13码': ean13,
                    'SKU': sku,
                    '生成数量': quantity,
                    '流水号': product_serial,
                    'RFID': rfid_code,
                    '创建时间': product.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })
                
                total_rfids += 1
        
        # 提交数据库事务
        db.session.commit()
        
        # 收集所有选中产品的款号，去重后用逗号隔开
        style_nos = set()
        for item in rfid_data:
            style_nos.add(item['款号'])
        style_no_str = ','.join(style_nos)
        
        # 生成Excel文件名，使用选择的款号
        filename = f"RFID_{style_no_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 生成Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "RFID生成结果"
        
        # 设置表头
        headers = list(rfid_data[0].keys()) if rfid_data else []
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据行
        for item in rfid_data:
            row = [item[header] for header in headers]
            ws.append(row)
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # 将Excel文件保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 返回Excel文件供下载
        from flask import send_file
        return send_file(output, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# 类别管理路由
@app.route('/category_management')
def category_management():
    """类别管理主页"""
    categories = Models.Category.query.order_by(Models.Category.name).all()
    subcategories = Models.SubCategory.query.order_by(Models.SubCategory.name).all()
    return render_template('category_management.html', categories=categories, subcategories=subcategories)

# 添加大类别路由
@app.route('/add_category', methods=['GET', 'POST'])
def add_category():
    """添加大类别"""
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().upper()
        
        # 检查名称和缩写是否已存在
        if Models.Category.query.filter_by(name=name).first():
            flash('大类别名称已存在！', 'danger')
        elif Models.Category.query.filter_by(code=code).first():
            flash('大类别缩写已存在！', 'danger')
        else:
            try:
                category = Models.Category(name=name, code=code)
                db.session.add(category)
                db.session.commit()
                flash('大类别添加成功！', 'success')
                return redirect(url_for('category_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'添加失败：{str(e)}', 'danger')
    
    return render_template('add_category.html')

# 编辑大类别路由
@app.route('/edit_category/<int:id>', methods=['GET', 'POST'])
def edit_category(id):
    """编辑大类别"""
    category = Models.Category.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().upper()
        
        # 检查名称和缩写是否已存在（排除当前记录）
        existing_name = Models.Category.query.filter(Models.Category.name == name, Models.Category.id != id).first()
        existing_code = Models.Category.query.filter(Models.Category.code == code, Models.Category.id != id).first()
        
        if existing_name:
            flash('大类别名称已存在！', 'danger')
        elif existing_code:
            flash('大类别缩写已存在！', 'danger')
        else:
            try:
                category.name = name
                category.code = code
                db.session.commit()
                flash('大类别更新成功！', 'success')
                return redirect(url_for('category_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'更新失败：{str(e)}', 'danger')
    
    return render_template('edit_category.html', category=category)

# 删除大类别路由
@app.route('/delete_category/<int:id>')
def delete_category(id):
    """删除大类别"""
    category = Models.Category.query.get_or_404(id)
    
    try:
        db.session.delete(category)
        db.session.commit()
        flash('大类别删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'danger')
    
    return redirect(url_for('category_management'))

# 颜色管理路由
@app.route('/color_management')
def color_management():
    """颜色管理主页"""
    colors = Models.Color.query.order_by(Models.Color.name).all()
    return render_template('color_management.html', colors=colors)

# 添加颜色路由
@app.route('/add_color', methods=['GET', 'POST'])
def add_color():
    """添加颜色"""
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().upper()
        
        # 检查名称和缩写是否已存在
        if Models.Color.query.filter_by(name=name).first():
            flash('颜色名称已存在！', 'danger')
        elif Models.Color.query.filter_by(code=code).first():
            flash('颜色缩写已存在！', 'danger')
        else:
            try:
                color = Models.Color(name=name, code=code)
                db.session.add(color)
                db.session.commit()
                flash('颜色添加成功！', 'success')
                return redirect(url_for('color_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'添加失败：{str(e)}', 'danger')
    
    return render_template('add_color.html')

# 编辑颜色路由
@app.route('/edit_color/<int:id>', methods=['GET', 'POST'])
def edit_color(id):
    """编辑颜色"""
    color = Models.Color.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().upper()
        
        # 检查名称和缩写是否已存在（排除当前记录）
        existing_name = Models.Color.query.filter(Models.Color.name == name, Models.Color.id != id).first()
        existing_code = Models.Color.query.filter(Models.Color.code == code, Models.Color.id != id).first()
        
        if existing_name:
            flash('颜色名称已存在！', 'danger')
        elif existing_code:
            flash('颜色缩写已存在！', 'danger')
        else:
            try:
                color.name = name
                color.code = code
                db.session.commit()
                flash('颜色更新成功！', 'success')
                return redirect(url_for('color_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'更新失败：{str(e)}', 'danger')
    
    return render_template('edit_color.html', color=color)

# 删除颜色路由
@app.route('/delete_color/<int:id>')
def delete_color(id):
    """删除颜色"""
    color = Models.Color.query.get_or_404(id)
    
    try:
        db.session.delete(color)
        db.session.commit()
        flash('颜色删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'danger')
    
    return redirect(url_for('color_management'))

# 添加小类别路由
@app.route('/add_subcategory', methods=['GET', 'POST'])
def add_subcategory():
    """添加小类别"""
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().upper()
        
        # 检查是否已存在相同名称或缩写
        existing_name = Models.SubCategory.query.filter_by(name=name).first()
        existing_code = Models.SubCategory.query.filter_by(code=code).first()
        
        if existing_name:
            flash('小类别名称已存在！', 'danger')
        elif existing_code:
            flash('小类别缩写已存在！', 'danger')
        else:
            try:
                subcategory = Models.SubCategory(name=name, code=code)
                db.session.add(subcategory)
                db.session.commit()
                flash('小类别添加成功！', 'success')
                return redirect(url_for('category_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'添加失败：{str(e)}', 'danger')
    
    return render_template('add_subcategory.html')

# 编辑小类别路由
@app.route('/edit_subcategory/<int:id>', methods=['GET', 'POST'])
def edit_subcategory(id):
    """编辑小类别"""
    subcategory = Models.SubCategory.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        code = request.form['code'].strip().upper()
        
        # 检查是否已存在相同名称或缩写（排除当前记录）
        existing_name = Models.SubCategory.query.filter(
            Models.SubCategory.name == name,
            Models.SubCategory.id != id
        ).first()
        existing_code = Models.SubCategory.query.filter(
            Models.SubCategory.code == code,
            Models.SubCategory.id != id
        ).first()
        
        if existing_name:
            flash('小类别名称已存在！', 'danger')
        elif existing_code:
            flash('小类别缩写已存在！', 'danger')
        else:
            try:
                subcategory.name = name
                subcategory.code = code
                db.session.commit()
                flash('小类别编辑成功！', 'success')
                return redirect(url_for('category_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'编辑失败：{str(e)}', 'danger')
    
    return render_template('edit_subcategory.html', subcategory=subcategory)

# 删除小类别路由
@app.route('/delete_subcategory/<int:id>')
def delete_subcategory(id):
    """删除小类别"""
    subcategory = Models.SubCategory.query.get_or_404(id)
    
    try:
        db.session.delete(subcategory)
        db.session.commit()
        flash('小类别删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'danger')
    
    return redirect(url_for('category_management'))

# 主函数
if __name__ == '__main__':
    # 初始化模型
    init_models()
    # 创建数据库表
    with app.app_context():
        db.create_all()
    # 运行应用
    app.run(debug=False, host='0.0.0.0', port=12456)